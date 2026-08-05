from __future__ import annotations

import csv
import json
import math
import os
import re
import shutil
import textwrap
import zipfile
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

ROOT = Path('/mnt/data')
SOURCE_WB = ROOT / 'AZ_BMS_Portfolio_Collision_Lollipop_Logo_Colors_v2_Backup.xlsx'
SOURCE_PNG = ROOT / 'AZ_BMS_Portfolio_Collision_Lollipop_Logo_Colors_v2.png'
SOURCE_PDF = ROOT / 'AZ_BMS_Portfolio_Collision_Lollipop_Logo_Colors_v2.pdf'
SOURCE_R = ROOT / 'build_az_bms_lollipop_logo_colors_v2.R'
SOURCE_PLOT_CSV = ROOT / 'AZ_BMS_Portfolio_Collision_Lollipop_Logo_Colors_v2_Plot_Data.csv'
OUT = ROOT / 'az_bms_portfolio_site'
ZIP_PATH = ROOT / 'AZ_BMS_Portfolio_Interactive_Website_GitHub_Pages.zip'

if OUT.exists():
    shutil.rmtree(OUT)
for sub in ['assets', 'data', 'downloads']:
    (OUT / sub).mkdir(parents=True, exist_ok=True)

AZ_PIPELINE = 'https://www.astrazeneca.com/content/dam/az/Investor_Relations/annual-report-2025/pdf/AstraZeneca_Development_Pipeline_2025.pdf'
BMS_PIPELINE = 'https://annual-report.bms.com/assets/bms-ar/documents/2025/2025-bms-development-portfolio.pdf'

# -----------------------------
# Load the matched pipeline data
# -----------------------------
wb_values = load_workbook(SOURCE_WB, data_only=True, read_only=True)
asset_ws = wb_values['Asset_Data']
asset_headers = [c.value for c in next(asset_ws.iter_rows(min_row=1, max_row=1))]
assets: list[dict[str, Any]] = []
for row in asset_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    rec = dict(zip(asset_headers, row))
    assets.append({
        'company': rec['Company'],
        'sourceTherapeuticArea': rec['Source therapeutic area'],
        'diseaseArea': rec['Normalized area'],
        'program': rec['Lead investigational asset / program'],
        'stage': rec['Highest disclosed phase in snapshot'],
        'modality': rec['Modality / mechanism'],
        'representativeIndications': rec['Representative area(s) under investigation'],
        'snapshotDate': rec['Source snapshot date'],
        'sourceUrl': rec['Source URL'],
        'notes': rec['Methodology notes'] or '',
    })

overlap_ws = wb_values['Overlap_Units']
overlap_units: list[dict[str, Any]] = []
for row in overlap_ws.iter_rows(min_row=5, values_only=True):
    if not row[0]:
        continue
    overlap_units.append({
        'diseaseArea': row[0],
        'factor': row[1],
        'unit': row[2],
        'azProgramsRaw': row[3] or '',
        'bmsProgramsRaw': row[4] or '',
        'directness': int(row[5] or 0),
        'rationale': row[6] or '',
        'azSourceUrl': row[7] or AZ_PIPELINE,
        'bmsSourceUrl': row[8] or BMS_PIPELINE,
    })

# Plot data is read from the prior export so counts match the published chart.
plot_rows: list[dict[str, Any]] = []
with SOURCE_PLOT_CSV.open(newline='', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        plot_rows.append({
            'diseaseArea': r['disease_area'],
            'plotLabel': r['plot_label'],
            'factor': r['factor'],
            'overlapCount': int(r['overlap_count']),
            'intensity': int(r['intensity']),
            'azPrograms': int(r['az_programs']),
            'bmsPrograms': int(r['bms_programs']),
            'relationship': r['relationship'],
        })

# -----------------------------
# Curated trial-level detail
# -----------------------------
# The launch window is an analyst scenario, calculated from the trial primary-completion
# year plus a stage-specific lag. It is intentionally not described as company guidance.
trial_details: dict[str, dict[str, Any]] = {
    'IPH5201': {
        'overlapIndication': 'NSCLC', 'trialId': 'NCT05742607', 'trialName': 'MATISSE',
        'trialStage': 'Phase II', 'potentialLaunchIndication': 'Resectable stage II-IIIA NSCLC',
        'clinicalSetting': 'Untreated; neoadjuvant therapy before surgery followed by adjuvant therapy',
        'primaryCompletion': '2025-06', 'lagLow': 3, 'lagHigh': 5,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT05742607',
        'detailNote': 'Early-stage study; no pivotal launch assumption is implied.'
    },
    'rilvegostomig': {
        'overlapIndication': 'NSCLC', 'trialId': 'NCT06692738', 'trialName': 'ARTEMIDE-Lung02',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Metastatic squamous NSCLC, PD-L1 >=1%',
        'clinicalSetting': 'First-line; rilvegostomig plus platinum chemotherapy',
        'primaryCompletion': '2029-02-05', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06692738',
        'detailNote': 'Representative pivotal study; the asset has additional NSCLC programs.'
    },
    'volrustomig': {
        'overlapIndication': 'NSCLC', 'trialId': 'NCT05984277', 'trialName': 'eVOLVE-Lung02',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Metastatic NSCLC, PD-L1 <50%',
        'clinicalSetting': 'First-line; volrustomig plus histology-specific chemotherapy',
        'primaryCompletion': '2028-04-24', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT05984277',
        'detailNote': 'Representative pivotal study.'
    },
    'pumitamig': {
        'overlapIndication': 'NSCLC', 'trialId': 'NCT07361510', 'trialName': 'ROSETTA Lung-202',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Advanced NSCLC, PD-L1 >=50%',
        'clinicalSetting': 'First-line monotherapy versus pembrolizumab',
        'primaryCompletion': '2031-10-14', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT07361510',
        'detailNote': 'Representative Phase III study; later timing reflects the currently posted trial schedule.'
    },
    'izalontamab brengitecan (iza-bren)': {
        'overlapIndication': 'NSCLC', 'trialId': 'NCT07100080', 'trialName': 'IZABRIGHT-Lung01',
        'trialStage': 'Phase II/III', 'potentialLaunchIndication': 'EGFR-mutated NSCLC after EGFR-TKI failure',
        'clinicalSetting': 'Post-EGFR TKI; compared with platinum-pemetrexed',
        'primaryCompletion': '2028-12-15', 'lagLow': 1, 'lagHigh': 3,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT07100080',
        'detailNote': 'Phase II/III design.'
    },
    'navlimetostat': {
        'overlapIndication': 'NSCLC', 'trialId': 'NCT06855771', 'trialName': 'MountainTAP-9',
        'trialStage': 'Phase II', 'potentialLaunchIndication': 'MTAP-deleted advanced or metastatic NSCLC',
        'clinicalSetting': 'Previously treated; after progression on prior therapies',
        'primaryCompletion': '2028-12-29', 'lagLow': 3, 'lagHigh': 5,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06855771',
        'detailNote': 'Phase II; a registrational path is not assumed.'
    },
    'nivolumab + relatlimab HD': {
        'overlapIndication': 'NSCLC', 'trialId': 'NCT06561386', 'trialName': 'RELATIVITY-1093',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Stage IV/recurrent nonsquamous NSCLC, PD-L1 >=1%',
        'clinicalSetting': 'First-line; nivolumab-relatlimab fixed dose plus chemotherapy',
        'primaryCompletion': '2030-07-30', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06561386',
        'detailNote': 'Potential label expansion of an approved combination; counted here as an investigational program in the matched snapshot.'
    },
    'FPI-2265': {
        'overlapIndication': 'Prostate cancer', 'trialId': 'NCT06402331', 'trialName': 'AlphaBreak',
        'trialStage': 'Phase II', 'potentialLaunchIndication': 'PSMA-positive metastatic castration-resistant prostate cancer',
        'clinicalSetting': 'Post-lutetium PSMA radioligand therapy; dose optimization',
        'primaryCompletion': '2026-12-23', 'lagLow': 3, 'lagHigh': 5,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06402331',
        'detailNote': 'Phase II; additional combination work is underway.'
    },
    'saruparib': {
        'overlapIndication': 'Prostate cancer', 'trialId': 'NCT06120491', 'trialName': 'EvoPAR-Prostate01',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Metastatic castration-sensitive prostate cancer',
        'clinicalSetting': 'First-line metastatic setting; with physician-choice new hormonal agent',
        'primaryCompletion': '2027-09-30', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06120491',
        'detailNote': 'Representative Phase III prostate study.'
    },
    'AR ligand-directed degrader': {
        'overlapIndication': 'Prostate cancer', 'trialId': 'NCT06764485', 'trialName': 'rechARge',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Metastatic castration-resistant prostate cancer',
        'clinicalSetting': 'After prior androgen-receptor pathway inhibitor; versus docetaxel or second ARPI',
        'primaryCompletion': '2027-09-12', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06764485',
        'detailNote': 'Program is also known as BMS-986365.'
    },
    'AZD0120': {
        'overlapIndication': 'Multiple myeloma', 'trialId': 'NCT05850234', 'trialName': 'DURGA-1',
        'trialStage': 'Phase Ib/II', 'potentialLaunchIndication': 'Relapsed/refractory multiple myeloma',
        'clinicalSetting': 'At least three prior lines; CD19/BCMA dual CAR-T',
        'primaryCompletion': '2027-08-31', 'lagLow': 4, 'lagHigh': 6,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT05850234',
        'detailNote': 'Early-stage CAR-T program; launch window is highly uncertain.'
    },
    'AZD0305': {
        'overlapIndication': 'Multiple myeloma', 'trialId': 'NCT06106945', 'trialName': 'Modular Phase I/II study',
        'trialStage': 'Phase I/II', 'potentialLaunchIndication': 'Multiple myeloma',
        'clinicalSetting': 'Monotherapy and combination dose escalation/expansion',
        'primaryCompletion': '2027-08-16', 'lagLow': 4, 'lagHigh': 6,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06106945',
        'detailNote': 'GPRC5D ADC; launch window is highly uncertain.'
    },
    'arlo-cel': {
        'overlapIndication': 'Multiple myeloma', 'trialId': 'NCT06615479', 'trialName': 'QUINTESSENTIAL-2',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Relapsed/refractory, lenalidomide-exposed multiple myeloma',
        'clinicalSetting': 'Relapsed/refractory setting; versus standard regimens',
        'primaryCompletion': '2027-12-30', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06615479',
        'detailNote': 'GPRC5D-directed CAR-T; program also has a later-line Phase II study.'
    },
    'iberdomide': {
        'overlapIndication': 'Multiple myeloma', 'trialId': 'NCT05827016', 'trialName': 'EXCALIBER-Maintenance',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Newly diagnosed multiple myeloma maintenance',
        'clinicalSetting': 'Maintenance following autologous stem-cell transplant',
        'primaryCompletion': '2029-03-15', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT05827016',
        'detailNote': 'Representative Phase III study; the asset has a broader development program.'
    },
    'mezigdomide': {
        'overlapIndication': 'Multiple myeloma', 'trialId': 'NCT05552976', 'trialName': 'SUCCESSOR-2',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Relapsed/refractory multiple myeloma',
        'clinicalSetting': 'Mezigdomide-carfilzomib-dexamethasone versus carfilzomib-dexamethasone',
        'primaryCompletion': '2026-07-18', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT05552976',
        'detailNote': 'BMS reported positive Phase III results in 2026; launch timing remains an analyst scenario.'
    },
    'surovatamig': {
        'overlapIndication': 'B-cell lymphoma', 'trialId': 'NCT06549595', 'trialName': 'SOUNDTRACK-F1',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Previously untreated follicular lymphoma',
        'clinicalSetting': 'First-line; surovatamig plus rituximab versus chemoimmunotherapy',
        'primaryCompletion': '2031-11-26', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06549595',
        'detailNote': 'Representative pivotal study in the B-cell lymphoma territory.'
    },
    'golcadomide': {
        'overlapIndication': 'B-cell lymphoma', 'trialId': 'NCT06356129', 'trialName': 'GOLSEEK-1',
        'trialStage': 'Phase III', 'potentialLaunchIndication': 'Previously untreated high-risk large B-cell lymphoma',
        'clinicalSetting': 'First-line; golcadomide plus R-CHOP',
        'primaryCompletion': '2028-08-14', 'lagLow': 1, 'lagHigh': 2,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06356129',
        'detailNote': 'Representative Phase III study; additional follicular lymphoma work is underway.'
    },
    'AZD5462': {
        'overlapIndication': 'Heart failure', 'trialId': 'NCT06299826', 'trialName': 'LUMINARA',
        'trialStage': 'Phase IIb', 'potentialLaunchIndication': 'Chronic heart failure',
        'clinicalSetting': 'Stable chronic heart failure; mechanism-focused proof of concept',
        'primaryCompletion': '2026-02-10', 'lagLow': 3, 'lagHigh': 5,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06299826',
        'detailNote': 'Phase IIb; no pivotal launch assumption is implied.'
    },
    'MYK-224': {
        'overlapIndication': 'Heart failure', 'trialId': 'NCT06122779', 'trialName': 'AURORA-HFpEF',
        'trialStage': 'Phase IIa', 'potentialLaunchIndication': 'Heart failure with preserved ejection fraction',
        'clinicalSetting': 'Symptomatic HFpEF; randomized placebo-controlled proof of concept',
        'primaryCompletion': '2026-07-03', 'lagLow': 3, 'lagHigh': 5,
        'trialUrl': 'https://clinicaltrials.gov/study/NCT06122779',
        'detailNote': 'Cardiac myosin inhibitor; not the same indication as approved Camzyos.'
    },
}

# Commercial context. Sales are not indication-specific unless explicitly stated.
commercial_benchmarks = [
    {
        'id': 'NSCLC_TAGRISSO_2025', 'diseaseArea': 'Solid-tumor oncology', 'overlapUnit': 'NSCLC',
        'product': 'Tagrisso', 'company': 'AstraZeneca', 'year': 2025,
        'us': 3064, 'emergingMarkets': 1971, 'europe': 1423, 'establishedRow': 796,
        'worldwideRevenue': 7254, 'basis': '2025 product sales; calculated from company-reported regions',
        'relevance': 'Commercial scale benchmark for a leading NSCLC franchise.',
        'limitations': 'Worldwide brand revenue across approved indications; not a forecast and not specific to the overlapping programs.',
        'sourceUrl': 'https://www.sec.gov/Archives/edgar/data/901832/000110465926019130/azn-20251231x20f.htm'
    },
    {
        'id': 'NSCLC_OPDIVO_2025', 'diseaseArea': 'Solid-tumor oncology', 'overlapUnit': 'NSCLC',
        'product': 'Opdivo', 'company': 'Bristol Myers Squibb', 'year': 2025,
        'us': None, 'emergingMarkets': None, 'europe': None, 'establishedRow': None,
        'worldwideRevenue': 10000, 'basis': '2025 worldwide revenue',
        'relevance': 'Commercial scale benchmark for BMS immuno-oncology.',
        'limitations': 'All approved indications, not NSCLC-specific; used as a scale benchmark rather than a direct analogue.',
        'sourceUrl': 'https://www.bms.com/investors/financial-reporting/key-facts.html'
    },
    {
        'id': 'MM_REVLIMID_2021', 'diseaseArea': 'Hematology', 'overlapUnit': 'Multiple myeloma',
        'product': 'Revlimid', 'company': 'Bristol Myers Squibb', 'year': 2021,
        'us': None, 'emergingMarkets': None, 'europe': None, 'establishedRow': None,
        'worldwideRevenue': 12821, 'basis': '2021 worldwide revenue; historical high-water benchmark',
        'relevance': 'Illustrates the historic commercial scale of the multiple-myeloma franchise.',
        'limitations': 'Brand revenue across indications; historical revenue does not predict future asset sales.',
        'sourceUrl': 'https://news.bms.com/news/details/2022/Bristol-Myers-Squibb-Reports-Fourth-Quarter-and-Full-Year-Financial-Results-for-2021/default.aspx'
    },
    {
        'id': 'MM_POMALYST_2021', 'diseaseArea': 'Hematology', 'overlapUnit': 'Multiple myeloma',
        'product': 'Pomalyst / Imnovid', 'company': 'Bristol Myers Squibb', 'year': 2021,
        'us': None, 'emergingMarkets': None, 'europe': None, 'establishedRow': None,
        'worldwideRevenue': 3332, 'basis': '2021 worldwide revenue',
        'relevance': 'Additional multiple-myeloma commercial benchmark.',
        'limitations': 'Brand revenue across indications; not a forecast.',
        'sourceUrl': 'https://news.bms.com/news/details/2022/Bristol-Myers-Squibb-Reports-Fourth-Quarter-and-Full-Year-Financial-Results-for-2021/default.aspx'
    },
    {
        'id': 'LYMPHOMA_BREYANZI_2025', 'diseaseArea': 'Hematology', 'overlapUnit': 'B-cell lymphoma',
        'product': 'Breyanzi', 'company': 'Bristol Myers Squibb', 'year': 2025,
        'us': None, 'emergingMarkets': None, 'europe': None, 'establishedRow': None,
        'worldwideRevenue': 1400, 'basis': '2025 worldwide revenue',
        'relevance': 'Commercial benchmark for an established B-cell lymphoma cell-therapy franchise.',
        'limitations': 'All approved indications; not a direct modality match for every program.',
        'sourceUrl': 'https://www.bms.com/investors/financial-reporting/key-facts.html'
    },
    {
        'id': 'HF_CAMZYOS_2025', 'diseaseArea': 'Cardiovascular', 'overlapUnit': 'Heart failure',
        'product': 'Camzyos', 'company': 'Bristol Myers Squibb', 'year': 2025,
        'us': None, 'emergingMarkets': None, 'europe': None, 'establishedRow': None,
        'worldwideRevenue': 1100, 'basis': '2025 worldwide revenue',
        'relevance': 'Commercial benchmark for a cardiac-myosin franchise.',
        'limitations': 'Approved for obstructive hypertrophic cardiomyopathy, not HFpEF; included only as modality/franchise context.',
        'sourceUrl': 'https://www.bms.com/investors/financial-reporting/key-facts.html'
    },
]

# Aliases connect shorthand evidence text to the canonical asset names.
aliases = {
    'iza-bren': 'izalontamab brengitecan (iza-bren)',
    'AR ligand-directed degrader': 'AR ligand-directed degrader',
    'torvutatug samrotecan': 'torvutatug samrotecan (AZD5335)',
    'nivolumab + relatlimab HD': 'nivolumab + relatlimab HD',
    'zola-cel': 'zola-cel (CD19 NEX-T)',
}

def slugify(s: str) -> str:
    s = s.lower().replace('&', ' and ')
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s

def split_programs(raw: str) -> list[str]:
    return [x.strip() for x in raw.split(';') if x and x.strip()]

asset_by_name: dict[str, dict[str, Any]] = {a['program']: a for a in assets}
for short, canonical in aliases.items():
    if canonical in asset_by_name:
        asset_by_name[short] = asset_by_name[canonical]

for u in overlap_units:
    u['azPrograms'] = split_programs(u.pop('azProgramsRaw'))
    u['bmsPrograms'] = split_programs(u.pop('bmsProgramsRaw'))
    u['id'] = f"{slugify(u['diseaseArea'])}--{slugify(u['factor'])}--{slugify(u['unit'])}"
    u['cellId'] = f"{slugify(u['diseaseArea'])}--{slugify(u['factor'])}"

# Enrich all asset records. Trial-specific fields are only present where they can be tied to a representative study.
enriched_programs: list[dict[str, Any]] = []
for a in assets:
    detail = trial_details.get(a['program'], {})
    primary = detail.get('primaryCompletion', '')
    year_match = re.match(r'^(\d{4})', primary or '')
    pc_year = int(year_match.group(1)) if year_match else None
    low = detail.get('lagLow')
    high = detail.get('lagHigh')
    launch_start = pc_year + low if pc_year and isinstance(low, int) else None
    launch_end = pc_year + high if pc_year and isinstance(high, int) else None
    launch_window = f"{launch_start}-{launch_end}" if launch_start and launch_end else 'Not estimated'
    enriched_programs.append({
        **a,
        'trialId': detail.get('trialId', ''),
        'trialName': detail.get('trialName', ''),
        'trialStage': detail.get('trialStage', a['stage']),
        'overlapIndication': detail.get('overlapIndication', ''),
        'potentialLaunchIndication': detail.get('potentialLaunchIndication', ''),
        'clinicalSetting': detail.get('clinicalSetting', ''),
        'primaryCompletion': primary,
        'launchLagLowYears': low,
        'launchLagHighYears': high,
        'illustrativeLaunchWindow': launch_window,
        'historicalSales': 'Not launched - no historical sales',
        'trialUrl': detail.get('trialUrl', ''),
        'detailNote': detail.get('detailNote', ''),
    })

# Cell-level plot data with stable IDs.
for r in plot_rows:
    r['diseaseSlug'] = slugify(r['diseaseArea'])
    r['factorSlug'] = slugify(r['factor'])
    r['cellId'] = f"{r['diseaseSlug']}--{r['factorSlug']}"

# -----------------------------
# CSV exports
# -----------------------------
def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    if not rows:
        return
    if fieldnames is None:
        fieldnames = list(rows[0].keys())
    with path.open('w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)

write_csv(OUT / 'data' / 'chart_data.csv', plot_rows)
write_csv(OUT / 'data' / 'overlap_units.csv', overlap_units)
write_csv(OUT / 'data' / 'program_details.csv', enriched_programs)
write_csv(OUT / 'data' / 'commercial_benchmarks.csv', commercial_benchmarks)

# -----------------------------
# Workbook: add website data sheets
# -----------------------------
WEB_WB = OUT / 'downloads' / 'AZ_BMS_Portfolio_Interactive_Website_Backup.xlsx'
shutil.copy2(SOURCE_WB, WEB_WB)
wb = load_workbook(WEB_WB)
for name in ['Website_README', 'Program_Details', 'Commercial_Benchmarks', 'Website_Export']:
    if name in wb.sheetnames:
        del wb[name]

# Style helpers
DARK = '172231'
AZ_PLUM = '8A1F68'
AZ_GOLD = 'F2C300'
BMS_NAVY = '283B75'
BMS_VIOLET = '7758A6'
RED = 'CE1F3A'
LIGHT = 'F5F1EB'
WHITE = 'FFFFFF'
GRAY = '667085'
GREEN = '008000'
BLUE = '0000FF'
ORANGE = 'F4B183'
THIN_GRAY = Side(style='thin', color='D9DEE7')

def setup_sheet(ws, freeze='A2'):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze

# Website_README
ws = wb.create_sheet('Website_README', 0)
setup_sheet(ws, freeze='A7')
ws.merge_cells('A1:H1')
ws['A1'] = 'Interactive website package - data and methodology'
ws['A1'].font = Font(name='Aptos Display', size=20, bold=True, color=WHITE)
ws['A1'].fill = PatternFill('solid', fgColor=DARK)
ws['A1'].alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 34
readme_rows = [
    ('Package purpose', 'Back-up data for the GitHub Pages website. The interactive chart links each overlap bubble to the exact overlap units and relevant programs.'),
    ('Matched portfolio snapshot', 'AstraZeneca 10 Feb 2026 and BMS 5 Feb 2026 pipeline disclosures.'),
    ('Late-stage program definition', 'Unique lead investigational asset in Phase II or Phase III/pivotal/registration development; approved-indication expansions and major lifecycle-management programs excluded from the core count.'),
    ('Overlap count definition', 'Number of distinct shared indication territories, modality classes, target/mechanism themes, or development-stage bands - not all possible AZ-BMS asset-pair permutations.'),
    ('Launch window', 'Illustrative analyst scenario calculated from trial primary-completion year plus a stage-specific lag. It is not company guidance, a probability-adjusted forecast, or a regulatory prediction.'),
    ('Historical sales', 'Investigational assets have no historical sales. The Commercial_Benchmarks sheet provides company-reported brand revenue for scale context; figures are generally not indication-specific.'),
    ('Website deployment', 'Upload the package contents to a GitHub repository and enable GitHub Pages from the main branch / root folder. No build step is required.'),
]
for i, (k, v) in enumerate(readme_rows, start=3):
    ws.cell(i, 1, k).font = Font(bold=True, color=DARK)
    ws.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    if i % 2:
        for c in range(1, 9):
            ws.cell(i, c).fill = PatternFill('solid', fgColor='F8F6F2')
    ws.row_dimensions[i].height = 42
ws.column_dimensions['A'].width = 28
for c in range(2, 9): ws.column_dimensions[get_column_letter(c)].width = 18

# Program_Details
ws = wb.create_sheet('Program_Details')
setup_sheet(ws)
program_headers = [
    'Disease area', 'Overlap indication', 'Company', 'Program', 'Pipeline stage', 'Representative modality / mechanism',
    'ClinicalTrials.gov ID', 'Representative trial', 'Potential launch indication (analyst scenario)', 'Clinical setting / line of therapy',
    'Primary completion date', 'Primary completion year', 'Launch lag - low (years)', 'Launch lag - high (years)',
    'Illustrative launch start', 'Illustrative launch end', 'Illustrative launch window', 'Historical sales of asset',
    'Commercial benchmark IDs', 'Trial URL', 'Pipeline source URL', 'Notes'
]
ws.append(program_headers)
benchmark_by_unit: dict[str, list[str]] = {}
for b in commercial_benchmarks:
    benchmark_by_unit.setdefault(b['overlapUnit'], []).append(b['id'])

curated_programs = [p for p in enriched_programs if p['trialId']]
curated_programs.sort(key=lambda x: (x['diseaseArea'], x['overlapIndication'], x['company'], x['program']))
for idx, p in enumerate(curated_programs, start=2):
    # Use an actual date when day precision exists; month-only values remain text to avoid false precision.
    pc = p['primaryCompletion']
    pc_value: Any = pc
    if re.match(r'^\d{4}-\d{2}-\d{2}$', pc):
        pc_value = datetime.strptime(pc, '%Y-%m-%d').date()
    row = [
        p['diseaseArea'], p['overlapIndication'], p['company'], p['program'], p['trialStage'], p['modality'],
        p['trialId'], p['trialName'], p['potentialLaunchIndication'], p['clinicalSetting'], pc_value,
        None, p['launchLagLowYears'], p['launchLagHighYears'], None, None, None,
        p['historicalSales'], '; '.join(benchmark_by_unit.get(p['overlapIndication'], [])), p['trialUrl'], p['sourceUrl'], p['detailNote']
    ]
    ws.append(row)
    # Formula-driven launch estimate.
    if isinstance(pc_value, date):
        ws.cell(idx, 12, f'=YEAR(K{idx})')
    else:
        ws.cell(idx, 12, f'=IF(K{idx}="","",VALUE(LEFT(K{idx},4)))')
    ws.cell(idx, 15, f'=IF(OR(L{idx}="",M{idx}=""),"",L{idx}+M{idx})')
    ws.cell(idx, 16, f'=IF(OR(L{idx}="",N{idx}=""),"",L{idx}+N{idx})')
    ws.cell(idx, 17, f'=IF(O{idx}="","Not estimated",TEXT(O{idx},"0")&"-"&TEXT(P{idx},"0"))')

# Header styling
for cell in ws[1]:
    cell.fill = PatternFill('solid', fgColor=DARK)
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[1].height = 50
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    company = row[2].value
    row[2].font = Font(color=AZ_PLUM if company == 'AstraZeneca' else BMS_NAVY, bold=True)
    # Imported/linked source fields green; formulas black; scenario controls blue.
    for col in [7, 8, 10, 11, 20, 21]:
        row[col-1].font = Font(color=GREEN)
    for col in [13, 14]:
        row[col-1].font = Font(color=BLUE)
    row[17].fill = PatternFill('solid', fgColor='FFF1F2')
    if row[10].is_date:
        row[10].number_format = 'yyyy-mm-dd'
widths = [24, 20, 20, 28, 16, 32, 18, 26, 34, 42, 19, 18, 18, 18, 18, 18, 22, 28, 28, 40, 44, 44]
for i, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.auto_filter.ref = f'A1:V{ws.max_row}'
tab = Table(displayName='ProgramDetailsTable', ref=f'A1:V{ws.max_row}')
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

# Commercial_Benchmarks
ws = wb.create_sheet('Commercial_Benchmarks')
setup_sheet(ws)
bench_headers = ['Benchmark ID', 'Disease area', 'Overlap indication', 'Product', 'Company', 'Year', 'US revenue ($m)', 'Emerging Markets revenue ($m)', 'Europe revenue ($m)', 'Established ROW revenue ($m)', 'Worldwide revenue ($m)', 'Sales basis', 'Relevance', 'Limitations', 'Source URL']
ws.append(bench_headers)
for i, b in enumerate(commercial_benchmarks, start=2):
    ws.append([b['id'], b['diseaseArea'], b['overlapUnit'], b['product'], b['company'], b['year'], b['us'], b['emergingMarkets'], b['europe'], b['establishedRow'], b['worldwideRevenue'], b['basis'], b['relevance'], b['limitations'], b['sourceUrl']])
    if b['id'] == 'NSCLC_TAGRISSO_2025':
        ws.cell(i, 11, f'=SUM(G{i}:J{i})')
for cell in ws[1]:
    cell.fill = PatternFill('solid', fgColor=DARK)
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[1].height = 44
for row in ws.iter_rows(min_row=2):
    for cell in row: cell.alignment = Alignment(wrap_text=True, vertical='top')
    for c in range(7, 12): row[c-1].number_format = '$#,##0;[Red]($#,##0);-'
    row[14].font = Font(color=GREEN)
    row[4].font = Font(color=AZ_PLUM if row[4].value == 'AstraZeneca' else BMS_NAVY, bold=True)
bench_widths = [28, 24, 20, 20, 22, 10, 18, 24, 18, 22, 22, 32, 42, 48, 52]
for i, width in enumerate(bench_widths, start=1): ws.column_dimensions[get_column_letter(i)].width = width
ws.auto_filter.ref = f'A1:O{ws.max_row}'
tab = Table(displayName='CommercialBenchmarksTable', ref=f'A1:O{ws.max_row}')
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium4', showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

# Website_Export: a clean machine-readable chart/cell table.
ws = wb.create_sheet('Website_Export')
setup_sheet(ws)
export_headers = ['Cell ID', 'Disease area', 'Plot label', 'Factor', 'Overlap count', 'Directness', 'AZ programs', 'BMS programs', 'Relationship', 'Overlap units']
ws.append(export_headers)
units_by_cell: dict[str, list[str]] = {}
for u in overlap_units:
    units_by_cell.setdefault(u['cellId'], []).append(u['unit'])
for r in plot_rows:
    ws.append([r['cellId'], r['diseaseArea'], r['plotLabel'], r['factor'], r['overlapCount'], r['intensity'], r['azPrograms'], r['bmsPrograms'], r['relationship'], '; '.join(units_by_cell.get(r['cellId'], []))])
for cell in ws[1]:
    cell.fill = PatternFill('solid', fgColor=DARK)
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(wrap_text=True, vertical='center')
for row in ws.iter_rows(min_row=2):
    for cell in row: cell.alignment = Alignment(wrap_text=True, vertical='top')
    row[4].fill = PatternFill('solid', fgColor='FFF1F2') if row[4].value else PatternFill(fill_type=None)
for i, width in enumerate([42, 30, 24, 28, 16, 14, 14, 14, 22, 48], start=1): ws.column_dimensions[get_column_letter(i)].width = width
ws.auto_filter.ref = f'A1:J{ws.max_row}'
tab = Table(displayName='WebsiteExportTable', ref=f'A1:J{ws.max_row}')
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

# Workbook calculation settings.
try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
except Exception:
    pass
wb.save(WEB_WB)

# -----------------------------
# Website assets/data
# -----------------------------
shutil.copy2(SOURCE_PNG, OUT / 'downloads' / 'AZ_BMS_Portfolio_Collision_Static.png')
shutil.copy2(SOURCE_PDF, OUT / 'downloads' / 'AZ_BMS_Portfolio_Collision_Static.pdf')
shutil.copy2(SOURCE_R, OUT / 'downloads' / 'build_az_bms_lollipop.R')
shutil.copy2(SOURCE_PLOT_CSV, OUT / 'downloads' / 'AZ_BMS_Portfolio_Collision_Plot_Data.csv')
shutil.copy2(SOURCE_PNG, OUT / 'assets' / 'figure-fallback.png')

site_data = {
    'meta': {
        'title': 'AstraZeneca + BMS: Portfolio Fit or Collision?',
        'asOf': '2026-08-05',
        'snapshot': 'Matched February 2026 company pipeline disclosures',
        'lateStageDefinition': 'Unique lead investigational asset in Phase II or Phase III/pivotal/registration development; approved-indication expansions and major lifecycle-management programs excluded.',
        'countDefinition': 'Distinct shared indication territories, modality classes, target/mechanism themes, or stage bands; not all possible asset-pair permutations.',
        'launchDisclaimer': 'Illustrative analyst scenario based on trial primary-completion timing and a stage-specific lag. Not company guidance or a probability-adjusted forecast.',
        'salesDisclaimer': 'Investigational assets have no historical sales. Commercial benchmarks are company-reported worldwide brand revenue and are generally not indication-specific.',
    },
    'palette': {
        'az': '#8A1F68', 'azAccent': '#F2C300', 'bms': '#283B75', 'bmsAccent': '#7758A6',
        'red1': '#F9D9D5', 'red2': '#F06A67', 'red3': '#CE1F3A', 'ink': '#172231', 'paper': '#F5F1EB'
    },
    'plot': plot_rows,
    'overlapUnits': overlap_units,
    'programs': enriched_programs,
    'commercialBenchmarks': commercial_benchmarks,
    'sources': [
        {'name': 'AstraZeneca Development Pipeline 2025', 'date': '2026-02-10', 'use': 'Matched portfolio count and program stage', 'url': AZ_PIPELINE},
        {'name': 'BMS Development Portfolio 2025', 'date': '2026-02-05', 'use': 'Matched portfolio count and program stage', 'url': BMS_PIPELINE},
        {'name': 'ClinicalTrials.gov', 'date': 'Records checked through 2026-08-05', 'use': 'Trial phase, setting and primary-completion timing', 'url': 'https://clinicaltrials.gov/'},
        {'name': 'AstraZeneca 2025 Form 20-F', 'date': '2026-02-24', 'use': 'Tagrisso commercial benchmark', 'url': 'https://www.sec.gov/Archives/edgar/data/901832/000110465926019130/azn-20251231x20f.htm'},
        {'name': 'BMS Key Facts', 'date': 'Updated May 2026', 'use': '2025 product revenue benchmarks', 'url': 'https://www.bms.com/investors/financial-reporting/key-facts.html'},
        {'name': 'BMS FY2021 Results', 'date': '2022-02-04', 'use': 'Historical multiple-myeloma franchise benchmark', 'url': 'https://news.bms.com/news/details/2022/Bristol-Myers-Squibb-Reports-Fourth-Quarter-and-Full-Year-Financial-Results-for-2021/default.aspx'},
    ]
}
with (OUT / 'assets' / 'data.js').open('w', encoding='utf-8') as f:
    f.write('window.PORTFOLIO_DATA = ')
    json.dump(site_data, f, ensure_ascii=False, separators=(',', ':'))
    f.write(';\n')
with (OUT / 'data' / 'portfolio_data.json').open('w', encoding='utf-8') as f:
    json.dump(site_data, f, ensure_ascii=False, indent=2)

# favicon: original abstract collision mark (no company logo assets).
(OUT / 'assets' / 'favicon.svg').write_text('''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64"><rect width="64" height="64" rx="14" fill="#172231"/><circle cx="25" cy="32" r="15" fill="#8A1F68"/><circle cx="39" cy="32" r="15" fill="#283B75"/><path d="M32 19a15 15 0 0 1 0 26 15 15 0 0 1 0-26Z" fill="#CE1F3A"/><circle cx="32" cy="32" r="4" fill="#F2C300"/></svg>''', encoding='utf-8')

# -----------------------------
# CSS
# -----------------------------
styles = r'''
:root {
  --paper:#f5f1eb; --card:#ffffff; --ink:#172231; --muted:#667085; --line:#d9dee7;
  --az:#8a1f68; --az-accent:#f2c300; --bms:#283b75; --bms-accent:#7758a6;
  --red1:#f9d9d5; --red2:#f06a67; --red3:#ce1f3a; --red-stroke:#9f1239;
  --shadow:0 18px 60px rgba(23,34,49,.10); --radius:24px;
}
*{box-sizing:border-box}
html{scroll-behavior:smooth}
body{margin:0;background:var(--paper);color:var(--ink);font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;line-height:1.5}
a{color:var(--bms);text-decoration-thickness:1px;text-underline-offset:3px}
a:hover{color:var(--az)}
button,input,select{font:inherit}
.skip-link{position:absolute;left:-10000px;top:auto;width:1px;height:1px;overflow:hidden}.skip-link:focus{left:16px;top:16px;width:auto;height:auto;background:#fff;padding:10px 14px;z-index:1000;border-radius:8px}
.site-header{position:sticky;top:0;z-index:20;background:rgba(245,241,235,.92);backdrop-filter:blur(18px);border-bottom:1px solid rgba(23,34,49,.08)}
.nav-wrap{max-width:1280px;margin:auto;padding:14px 28px;display:flex;gap:24px;align-items:center;justify-content:space-between}
.brand{display:flex;align-items:center;gap:11px;font-weight:800;letter-spacing:-.02em}.brand-mark{width:34px;height:34px;border-radius:10px;background:linear-gradient(90deg,var(--az) 0 44%,var(--red3) 44% 56%,var(--bms) 56%);box-shadow:inset 0 0 0 1px rgba(255,255,255,.35)}
.nav-links{display:flex;align-items:center;gap:22px;font-size:14px}.nav-links a{text-decoration:none;color:var(--ink);font-weight:650}.nav-links a:hover{color:var(--az)}
.container{max-width:1280px;margin:auto;padding:0 28px}
.hero{padding:72px 0 34px}.eyebrow{font-size:13px;letter-spacing:.11em;text-transform:uppercase;font-weight:800;color:var(--az);margin-bottom:14px}.hero h1{font-size:clamp(42px,6vw,76px);line-height:1.03;letter-spacing:-.045em;margin:0;max-width:1000px}.hero h1 .highlight{display:inline-block;color:var(--red3);background:#f8dfe5;border-radius:8px;padding:0 .12em .07em}.hero-copy{font-size:clamp(17px,2vw,22px);color:#344054;max-width:930px;margin:22px 0 0}.hero-actions{display:flex;gap:12px;flex-wrap:wrap;margin-top:26px}.button{display:inline-flex;align-items:center;justify-content:center;gap:8px;border-radius:999px;padding:11px 17px;font-weight:750;text-decoration:none;border:1px solid transparent;cursor:pointer}.button.primary{background:var(--ink);color:#fff}.button.secondary{background:#fff;color:var(--ink);border-color:var(--line)}.button:hover{transform:translateY(-1px);box-shadow:0 8px 24px rgba(23,34,49,.12)}
.status-note{display:inline-flex;gap:10px;align-items:flex-start;margin-top:24px;color:#475467;font-size:14px;background:rgba(255,255,255,.58);border:1px solid rgba(23,34,49,.08);padding:11px 14px;border-radius:12px}.status-dot{width:9px;height:9px;border-radius:50%;background:var(--red3);margin-top:6px;flex:0 0 auto}
.section{padding:34px 0}.section-head{display:flex;align-items:end;justify-content:space-between;gap:24px;margin-bottom:18px}.section-head h2{font-size:clamp(28px,4vw,46px);line-height:1.08;letter-spacing:-.035em;margin:0}.section-head p{max-width:660px;margin:8px 0 0;color:#475467}.kicker{color:var(--az);text-transform:uppercase;font-weight:800;letter-spacing:.09em;font-size:12px;margin-bottom:8px}
.card{background:var(--card);border:1px solid rgba(23,34,49,.08);border-radius:var(--radius);box-shadow:var(--shadow)}
.chart-card{padding:24px 22px 18px;overflow:hidden}.chart-toolbar{display:flex;justify-content:space-between;align-items:center;gap:18px;margin-bottom:10px}.chart-title{font-weight:850}.chart-subtitle{font-size:13px;color:var(--muted);margin-top:3px}.chart-hint{font-size:13px;color:#475467;background:#f7f8fa;border:1px solid var(--line);padding:8px 11px;border-radius:999px;white-space:nowrap}.chart-scroll{overflow-x:auto;overscroll-behavior-inline:contain;padding-bottom:6px}.chart-wrap{min-width:1040px;position:relative}.chart-wrap svg{width:100%;height:auto;display:block}.chart-wrap svg .bubble{cursor:pointer;transition:transform .16s ease,filter .16s ease;transform-box:fill-box;transform-origin:center}.chart-wrap svg .bubble:hover,.chart-wrap svg .bubble:focus{transform:scale(1.09);filter:drop-shadow(0 5px 8px rgba(159,18,57,.20));outline:none}.chart-wrap svg .bubble.selected circle{stroke:#172231;stroke-width:3}.chart-caption{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:12px;padding-top:14px;border-top:1px solid #edf0f4;color:#475467;font-size:13px}.legend-line{display:flex;align-items:center;gap:12px;flex-wrap:wrap}.legend-company{display:inline-flex;align-items:center;gap:7px}.legend-swatch{width:18px;height:12px;border-radius:3px;position:relative;overflow:hidden}.legend-swatch.az{background:var(--az);border-left:3px solid var(--az-accent)}.legend-swatch.bms{background:var(--bms);border-left:3px solid var(--bms-accent)}
.tooltip{position:fixed;z-index:50;pointer-events:none;background:var(--ink);color:#fff;padding:10px 12px;border-radius:10px;font-size:12px;max-width:250px;box-shadow:0 12px 32px rgba(0,0,0,.22);opacity:0;transform:translateY(4px);transition:opacity .12s,transform .12s}.tooltip.show{opacity:1;transform:none}.tooltip strong{display:block;font-size:13px;margin-bottom:2px}
.insight-strip{margin-top:18px;background:var(--ink);color:#fff;padding:26px 28px;border-radius:20px;display:grid;grid-template-columns:170px 1fr;gap:24px}.insight-label{font-size:12px;text-transform:uppercase;letter-spacing:.1em;color:#fac4d4;font-weight:800}.insight-strip strong{font-size:22px;line-height:1.25}.insight-strip p{color:#d6dce4;margin:8px 0 0}
.detail-card{padding:26px}.selection-head{display:flex;align-items:flex-start;justify-content:space-between;gap:18px;padding-bottom:20px;border-bottom:1px solid #edf0f4}.selection-head h3{font-size:28px;letter-spacing:-.025em;margin:2px 0 4px}.selection-meta{color:#667085;font-size:14px}.selection-badge{white-space:nowrap;border-radius:999px;background:#fff1f2;color:#9f1239;border:1px solid #fecdd3;font-weight:800;font-size:13px;padding:8px 12px}.unit-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:14px;margin:20px 0}.unit-card{border:1px solid var(--line);border-radius:16px;padding:16px;background:#fff;cursor:pointer;text-align:left;transition:.15s}.unit-card:hover,.unit-card.active{border-color:var(--red3);box-shadow:0 8px 24px rgba(206,31,58,.10);transform:translateY(-1px)}.unit-card.active{background:#fff7f8}.unit-top{display:flex;justify-content:space-between;align-items:flex-start;gap:12px}.unit-name{font-weight:850;font-size:17px}.directness{font-size:11px;text-transform:uppercase;letter-spacing:.06em;font-weight:850;border-radius:999px;padding:5px 8px}.directness.d1{background:var(--red1);color:#8c2636}.directness.d2{background:#fee2e2;color:#b42318}.directness.d3{background:var(--red3);color:#fff}.unit-card p{font-size:13px;color:#475467;margin:9px 0 12px}.program-chips{display:flex;flex-wrap:wrap;gap:6px}.chip{font-size:11px;border-radius:999px;padding:4px 7px;background:#f2f4f7;color:#344054}.chip.az{box-shadow:inset 3px 0 var(--az)}.chip.bms{box-shadow:inset 3px 0 var(--bms)}
.detail-controls{display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap;margin:24px 0 12px}.detail-controls h4{font-size:18px;margin:0}.search{min-width:260px;border:1px solid var(--line);background:#fff;border-radius:12px;padding:9px 12px;color:var(--ink)}
.table-wrap{overflow:auto;border:1px solid var(--line);border-radius:15px}.data-table{border-collapse:separate;border-spacing:0;width:100%;min-width:1320px;font-size:13px}.data-table th{position:sticky;top:0;background:var(--ink);color:#fff;text-align:left;padding:11px 12px;font-size:11px;text-transform:uppercase;letter-spacing:.045em;z-index:2}.data-table td{padding:12px;border-bottom:1px solid #edf0f4;vertical-align:top}.data-table tbody tr:nth-child(even){background:#fafbfc}.data-table tbody tr:hover{background:#fff8fa}.company-pill{display:inline-flex;border-radius:999px;padding:4px 8px;font-weight:800;font-size:11px;color:#fff}.company-pill.az{background:var(--az)}.company-pill.bms{background:var(--bms)}.stage-pill{display:inline-flex;border:1px solid #ccd3de;border-radius:999px;padding:3px 7px;white-space:nowrap;background:#fff}.muted{color:#667085}.nowrap{white-space:nowrap}.empty-state{padding:30px;text-align:center;color:#667085}.source-link{font-weight:750;white-space:nowrap}
.commercial-section{margin-top:24px}.commercial-section h4{font-size:18px;margin:0 0 10px}.benchmark-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:12px}.benchmark{border-radius:16px;padding:16px;background:#f8f7f4;border:1px solid #e5e2dc}.benchmark .value{font-size:27px;font-weight:900;letter-spacing:-.03em}.benchmark .label{font-weight:800}.benchmark .basis{color:#667085;font-size:12px;margin-top:4px}.benchmark .note{font-size:12px;color:#475467;margin-top:10px}.benchmark a{font-size:12px;font-weight:750}
.method-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}.method-card{padding:20px}.method-card .number{width:34px;height:34px;display:grid;place-items:center;border-radius:10px;background:var(--ink);color:#fff;font-weight:900;margin-bottom:13px}.method-card h3{margin:0 0 8px;font-size:18px}.method-card p{margin:0;color:#475467;font-size:14px}
.download-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:14px}.download-card{padding:19px;text-decoration:none;color:var(--ink);transition:.15s}.download-card:hover{transform:translateY(-2px);border-color:var(--bms)}.download-type{font-size:11px;text-transform:uppercase;letter-spacing:.08em;font-weight:850;color:var(--az)}.download-card h3{margin:7px 0 5px;font-size:18px}.download-card p{margin:0;color:#667085;font-size:13px}
.sources{padding:20px}.source-list{margin:0;padding:0;list-style:none}.source-list li{display:grid;grid-template-columns:minmax(180px,1.2fr) minmax(120px,.7fr) minmax(260px,2fr);gap:16px;padding:12px 0;border-bottom:1px solid #edf0f4;font-size:13px}.source-list li:last-child{border-bottom:0}.source-name{font-weight:800}.source-use{color:#667085}
footer{padding:38px 0 58px;color:#667085;font-size:13px}.footer-grid{display:flex;justify-content:space-between;gap:30px;align-items:flex-start;border-top:1px solid rgba(23,34,49,.12);padding-top:22px}.footer-grid strong{color:var(--ink)}
@media(max-width:900px){.nav-links{display:none}.hero{padding-top:48px}.insight-strip{grid-template-columns:1fr}.method-grid{grid-template-columns:1fr}.source-list li{grid-template-columns:1fr;gap:4px}.chart-caption{grid-template-columns:1fr}.selection-head{flex-direction:column}.container{padding:0 18px}.nav-wrap{padding:12px 18px}}
@media(max-width:600px){.hero h1{font-size:43px}.chart-card{padding:16px 10px}.detail-card{padding:18px 12px}.section{padding:28px 0}.search{width:100%;min-width:0}.footer-grid{flex-direction:column}}
@media(prefers-reduced-motion:reduce){html{scroll-behavior:auto}*{transition:none!important}}
'''
(OUT / 'assets' / 'styles.css').write_text(styles, encoding='utf-8')

# -----------------------------
# JavaScript
# -----------------------------
app_js = r'''
(() => {
  'use strict';
  const D = window.PORTFOLIO_DATA;
  const factors = ['Shared indications','Shared modalities','Shared target / mechanism','Stage alignment'];
  const diseaseOrder = [...new Map(D.plot.map(d => [d.diseaseArea, d])).values()];
  const unitsByCell = new Map();
  D.overlapUnits.forEach(u => {
    if (!unitsByCell.has(u.cellId)) unitsByCell.set(u.cellId, []);
    unitsByCell.get(u.cellId).push(u);
  });
  const programsByName = new Map(D.programs.map(p => [p.program, p]));
  const aliasMap = new Map([
    ['iza-bren','izalontamab brengitecan (iza-bren)'],
    ['torvutatug samrotecan','torvutatug samrotecan (AZD5335)'],
    ['zola-cel','zola-cel (CD19 NEX-T)']
  ]);
  const chartEl = document.getElementById('chart');
  const tooltip = document.getElementById('tooltip');
  const unitGrid = document.getElementById('unit-grid');
  const detailTitle = document.getElementById('detail-title');
  const detailMeta = document.getElementById('detail-meta');
  const detailBadge = document.getElementById('detail-badge');
  const tableBody = document.getElementById('program-table-body');
  const searchInput = document.getElementById('program-search');
  const benchmarkGrid = document.getElementById('benchmark-grid');
  const benchmarkSection = document.getElementById('commercial-section');
  let selectedCell = null;
  let selectedUnitId = null;
  let currentPrograms = [];

  const esc = s => String(s ?? '').replace(/[&<>'"]/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[c]));
  const slug = s => String(s).toLowerCase().replace(/&/g,' and ').replace(/[^a-z0-9]+/g,'-').replace(/^-|-$/g,'');
  const formatRevenue = m => m >= 1000 ? `$${(m/1000).toFixed(m % 1000 ? 1 : 0)}B` : `$${m}M`;
  const intensityLabel = i => ['None','Limited','Moderate','Strong'][i] || 'None';
  const factorShort = f => ({'Shared indications':'Indications','Shared modalities':'Modalities','Shared target / mechanism':'Target / mechanism','Stage alignment':'Stage'})[f] || f;
  const companyClass = c => c === 'AstraZeneca' ? 'az' : 'bms';
  const canonical = name => aliasMap.get(name) || name;

  function svgEl(name, attrs={}, text='') {
    const el = document.createElementNS('http://www.w3.org/2000/svg', name);
    Object.entries(attrs).forEach(([k,v]) => el.setAttribute(k, v));
    if (text !== '') el.textContent = text;
    return el;
  }
  function addMultilineText(svg, lines, x, y, attrs={}) {
    const t = svgEl('text', {x,y,...attrs});
    lines.forEach((line,i) => {
      const sp = svgEl('tspan', {x, dy: i===0 ? 0 : 14}, line);
      t.appendChild(sp);
    });
    svg.appendChild(t); return t;
  }
  function wrapLabel(label) {
    const map = {
      'Solid tumors':['Solid','tumors'], 'Hematology':['Hematology'],
      'Immunology / resp. / fibrosis':['Immunology /','resp. / fibrosis'], 'CV':['CV'],
      'Rare disease':['Rare','disease'], 'Renal / metabolic':['Renal /','metabolic'],
      'Infectious disease':['Infectious','disease'], 'Neuroscience':['Neuroscience']
    };
    return map[label] || [label];
  }

  function renderChart() {
    const W=1120,H=720,left=180,right=35,top=38,barBase=285,matrixTop=360,matrixGap=83;
    const colW=(W-left-right)/diseaseOrder.length;
    const svg=svgEl('svg',{viewBox:`0 0 ${W} ${H}`,role:'img','aria-labelledby':'chart-svg-title chart-svg-desc'});
    svg.appendChild(svgEl('title',{id:'chart-svg-title'},'AstraZeneca and BMS portfolio collision matrix'));
    svg.appendChild(svgEl('desc',{id:'chart-svg-desc'},'Stacked bars show late-stage program counts by disease area. Clickable red bubbles show the count and directness of shared indication, modality, target and stage dimensions.'));

    svg.appendChild(svgEl('text',{x:18,y:24,'font-size':13,'font-weight':850,fill:D.palette.ink,'letter-spacing':'.04em'},'LATE-STAGE PROGRAM FOOTPRINT'));
    svg.appendChild(svgEl('text',{x:18,y:45,'font-size':12,fill:'#667085'},'Unique lead Phase II or Phase III/pivotal/registration assets/programs'));

    // Company legend
    const lx=W-270,ly=22;
    const azRect=svgEl('rect',{x:lx,y:ly-10,width:22,height:13,rx:2,fill:D.palette.az}); svg.appendChild(azRect);
    svg.appendChild(svgEl('rect',{x:lx,y:ly-10,width:3,height:13,fill:D.palette.azAccent}));
    svg.appendChild(svgEl('text',{x:lx+29,y:ly+1,'font-size':12,fill:'#344054'},'AstraZeneca'));
    svg.appendChild(svgEl('rect',{x:lx+125,y:ly-10,width:22,height:13,rx:2,fill:D.palette.bms}));
    svg.appendChild(svgEl('rect',{x:lx+125,y:ly-10,width:3,height:13,fill:D.palette.bmsAccent}));
    svg.appendChild(svgEl('text',{x:lx+154,y:ly+1,'font-size':12,fill:'#344054'},'BMS'));

    const maxTotal=Math.max(...diseaseOrder.map(d=>d.azPrograms+d.bmsPrograms));
    const barScale=165/maxTotal;
    svg.appendChild(svgEl('line',{x1:left-5,y1:barBase,x2:W-right,y2:barBase,stroke:'#d9dee7','stroke-width':1.2}));
    diseaseOrder.forEach((d,i)=>{
      const x=left+i*colW+colW/2, bw=61;
      const azH=d.azPrograms*barScale,bmsH=d.bmsPrograms*barScale,total=d.azPrograms+d.bmsPrograms;
      if (d.azPrograms>0){
        svg.appendChild(svgEl('rect',{x:x-bw/2,y:barBase-azH,width:bw,height:azH,fill:D.palette.az}));
        svg.appendChild(svgEl('rect',{x:x-bw/2,y:barBase-azH,width:3,height:azH,fill:D.palette.azAccent}));
        if(azH>25) svg.appendChild(svgEl('text',{x,y:barBase-azH/2+4,'text-anchor':'middle','font-size':13,'font-weight':800,fill:'#fff'},String(d.azPrograms)));
      }
      if (d.bmsPrograms>0){
        svg.appendChild(svgEl('rect',{x:x-bw/2,y:barBase-azH-bmsH,width:bw,height:bmsH,fill:D.palette.bms}));
        svg.appendChild(svgEl('rect',{x:x-bw/2,y:barBase-azH-bmsH,width:3,height:bmsH,fill:D.palette.bmsAccent}));
        if(bmsH>25) svg.appendChild(svgEl('text',{x,y:barBase-azH-bmsH/2+4,'text-anchor':'middle','font-size':13,'font-weight':800,fill:'#fff'},String(d.bmsPrograms)));
      }
      svg.appendChild(svgEl('text',{x,y:barBase-azH-bmsH-9,'text-anchor':'middle','font-size':15,'font-weight':900,fill:D.palette.ink},String(total)));
      addMultilineText(svg,wrapLabel(d.plotLabel),x,barBase+24,{'text-anchor':'middle','font-size':12,'font-weight':760,fill:D.palette.ink});
      svg.appendChild(svgEl('line',{x1:x,y1:matrixTop-42,x2:x,y2:matrixTop+(factors.length-1)*matrixGap+48,stroke:'#ece7e1','stroke-width':1}));
    });

    factors.forEach((f,fi)=>{
      const y=matrixTop+fi*matrixGap;
      svg.appendChild(svgEl('line',{x1:left-5,y1:y,x2:W-right,y2:y,stroke:'#ddd7cf','stroke-width':1.1}));
      svg.appendChild(svgEl('text',{x:left-15,y:y+5,'text-anchor':'end','font-size':13,fill:D.palette.ink},f));
    });

    const radius = count => count===0?7:count===1?13:count===2?16:count===3?19:22;
    diseaseOrder.forEach((d,i)=>{
      const x=left+i*colW+colW/2;
      const nonzero=[];
      factors.forEach((f,fi)=>{
        const y=matrixTop+fi*matrixGap;
        const r=D.plot.find(p=>p.diseaseArea===d.diseaseArea&&p.factor===f);
        if(!r) return;
        if(r.overlapCount>0) nonzero.push(y);
      });
      if(nonzero.length>1) svg.appendChild(svgEl('line',{x1:x,y1:Math.min(...nonzero),x2:x,y2:Math.max(...nonzero),stroke:'#344054','stroke-width':2}));
      factors.forEach((f,fi)=>{
        const y=matrixTop+fi*matrixGap;
        const r=D.plot.find(p=>p.diseaseArea===d.diseaseArea&&p.factor===f);
        if(!r) return;
        if(r.overlapCount===0){
          svg.appendChild(svgEl('circle',{cx:x,cy:y,r:7,fill:'#fff',stroke:'#bfc7d3','stroke-width':1.5}));
          return;
        }
        const g=svgEl('g',{class:'bubble',tabindex:'0',role:'button','data-cell-id':r.cellId,'aria-label':`${d.plotLabel}, ${f}: ${r.overlapCount} overlap units, ${intensityLabel(r.intensity)} directness`});
        const fill=[null,D.palette.red1,D.palette.red2,D.palette.red3][r.intensity];
        const cr=radius(r.overlapCount);
        g.appendChild(svgEl('circle',{cx:x,cy:y,r:cr,fill,stroke:'#9f1239','stroke-width':1.5}));
        g.appendChild(svgEl('text',{x,y:y+5,'text-anchor':'middle','font-size':13,'font-weight':900,fill:r.intensity===3?'#fff':'#7f1d2d'},String(r.overlapCount)));
        const activate=()=>selectCell(r.cellId,true);
        g.addEventListener('click',activate);
        g.addEventListener('keydown',e=>{if(e.key==='Enter'||e.key===' '){e.preventDefault();activate();}});
        g.addEventListener('pointerenter',e=>showTooltip(e,r));
        g.addEventListener('pointermove',moveTooltip);
        g.addEventListener('pointerleave',hideTooltip);
        svg.appendChild(g);
      });
    });

    // Legends
    const ly2=690;
    svg.appendChild(svgEl('text',{x:18,y:ly2-22,'font-size':11,'font-weight':850,fill:'#344054'},'BUBBLE SIZE = OVERLAP COUNT'));
    [1,2,4].forEach((n,j)=>{
      const x=220+j*95, rr=radius(n);
      svg.appendChild(svgEl('circle',{cx:x,cy:ly2-25,r:rr,fill:'#fff',stroke:'#475467','stroke-width':1.3}));
      svg.appendChild(svgEl('text',{x,y:ly2-21,'text-anchor':'middle','font-size':11,'font-weight':850,fill:'#172231'},String(n)));
    });
    svg.appendChild(svgEl('text',{x:585,y:ly2-22,'font-size':11,'font-weight':850,fill:'#344054'},'RED TONE = DIRECTNESS'));
    [1,2,3].forEach((n,j)=>{
      const x=770+j*120,fill=[null,D.palette.red1,D.palette.red2,D.palette.red3][n];
      svg.appendChild(svgEl('circle',{cx:x,cy:ly2-25,r:12,fill,stroke:'#9f1239','stroke-width':1.2}));
      svg.appendChild(svgEl('text',{x:x+21,y:ly2-21,'font-size':11,fill:'#344054'},intensityLabel(n).toLowerCase()));
    });
    chartEl.innerHTML=''; chartEl.appendChild(svg);
  }

  function showTooltip(e,r){
    tooltip.innerHTML=`<strong>${esc(r.plotLabel)} · ${esc(factorShort(r.factor))}</strong>${r.overlapCount} distinct overlap ${r.overlapCount===1?'unit':'units'} · ${esc(intensityLabel(r.intensity))} directness<br><span style="color:#cdd5df">Click for evidence and program details</span>`;
    tooltip.classList.add('show'); moveTooltip(e);
  }
  function moveTooltip(e){tooltip.style.left=`${Math.min(e.clientX+16,window.innerWidth-270)}px`;tooltip.style.top=`${Math.min(e.clientY+16,window.innerHeight-120)}px`;}
  function hideTooltip(){tooltip.classList.remove('show');}

  function getProgramsForUnits(units){
    const names=[];
    units.forEach(u=>[...u.azPrograms,...u.bmsPrograms].forEach(n=>names.push(canonical(n))));
    return [...new Set(names)].map(n=>programsByName.get(n)).filter(Boolean);
  }
  function renderUnitCards(units){
    if(!units.length){unitGrid.innerHTML='<div class="empty-state">No distinct overlap units were coded for this chart cell.</div>';return;}
    unitGrid.innerHTML=units.map(u=>`
      <button class="unit-card ${selectedUnitId===u.id?'active':''}" data-unit-id="${esc(u.id)}" type="button">
        <div class="unit-top"><span class="unit-name">${esc(u.unit)}</span><span class="directness d${u.directness}">${esc(intensityLabel(u.directness))}</span></div>
        <p>${esc(u.rationale)}</p>
        <div class="program-chips">
          ${u.azPrograms.map(p=>`<span class="chip az">AZ · ${esc(p)}</span>`).join('')}
          ${u.bmsPrograms.map(p=>`<span class="chip bms">BMS · ${esc(p)}</span>`).join('')}
        </div>
      </button>`).join('');
    unitGrid.querySelectorAll('.unit-card').forEach(btn=>btn.addEventListener('click',()=>selectUnit(btn.dataset.unitId)));
  }
  function selectUnit(id){
    selectedUnitId = selectedUnitId===id ? null : id;
    const units=unitsByCell.get(selectedCell)||[];
    renderUnitCards(units);
    currentPrograms=getProgramsForUnits(selectedUnitId?units.filter(u=>u.id===selectedUnitId):units);
    renderPrograms(); renderBenchmarks(selectedUnitId?units.filter(u=>u.id===selectedUnitId):units);
  }
  function renderPrograms(){
    const q=searchInput.value.trim().toLowerCase();
    const rows=currentPrograms.filter(p=>!q||[p.company,p.program,p.stage,p.modality,p.representativeIndications,p.overlapIndication,p.clinicalSetting,p.trialId].join(' ').toLowerCase().includes(q));
    if(!rows.length){tableBody.innerHTML='<tr><td colspan="10"><div class="empty-state">No matching program details. Clear the search or select another overlap unit.</div></td></tr>';return;}
    rows.sort((a,b)=>(a.company.localeCompare(b.company)||a.program.localeCompare(b.program)));
    tableBody.innerHTML=rows.map(p=>`
      <tr>
        <td><span class="company-pill ${companyClass(p.company)}">${p.company==='AstraZeneca'?'AZ':'BMS'}</span></td>
        <td><strong>${esc(p.program)}</strong><div class="muted">${esc(p.modality)}</div></td>
        <td><span class="stage-pill">${esc(p.trialStage||p.stage)}</span></td>
        <td>${p.trialId?`<strong>${esc(p.trialName)}</strong><br><a class="source-link" href="${esc(p.trialUrl)}" target="_blank" rel="noopener">${esc(p.trialId)} ↗</a>`:`<span class="muted">Pipeline snapshot only</span>`}</td>
        <td>${esc(p.potentialLaunchIndication||p.representativeIndications)}</td>
        <td>${esc(p.clinicalSetting||'See pipeline source')}</td>
        <td class="nowrap">${esc(p.primaryCompletion||'Not mapped')}</td>
        <td class="nowrap">${esc(p.illustrativeLaunchWindow||'Not estimated')}${p.primaryCompletion?'<div class="muted">illustrative</div>':''}</td>
        <td>${esc(p.historicalSales)}</td>
        <td><a class="source-link" href="${esc(p.sourceUrl)}" target="_blank" rel="noopener">Pipeline ↗</a>${p.detailNote?`<div class="muted">${esc(p.detailNote)}</div>`:''}</td>
      </tr>`).join('');
  }
  function renderBenchmarks(units){
    const unitNames=new Set(units.map(u=>u.unit));
    let benchmarks=D.commercialBenchmarks.filter(b=>unitNames.has(b.overlapUnit));
    if(!benchmarks.length && selectedCell){
      const area=(D.plot.find(p=>p.cellId===selectedCell)||{}).diseaseArea;
      benchmarks=D.commercialBenchmarks.filter(b=>b.diseaseArea===area);
    }
    benchmarkSection.hidden=!benchmarks.length;
    benchmarkGrid.innerHTML=benchmarks.map(b=>`
      <div class="benchmark">
        <div class="label">${esc(b.product)} · ${esc(String(b.year))}</div>
        <div class="value">${formatRevenue(b.worldwideRevenue)}</div>
        <div class="basis">${esc(b.basis)}</div>
        <div class="note">${esc(b.relevance)} ${esc(b.limitations)}</div>
        <a href="${esc(b.sourceUrl)}" target="_blank" rel="noopener">Company source ↗</a>
      </div>`).join('');
  }
  function updateSelectedBubble(){
    document.querySelectorAll('.bubble').forEach(b=>b.classList.toggle('selected',b.dataset.cellId===selectedCell));
  }
  function selectCell(cellId,scroll=false){
    hideTooltip();
    selectedCell=cellId; selectedUnitId=null; searchInput.value='';
    const r=D.plot.find(p=>p.cellId===cellId);
    if(!r)return;
    const units=unitsByCell.get(cellId)||[];
    detailTitle.textContent=`${r.plotLabel} · ${factorShort(r.factor)}`;
    detailMeta.textContent=units.length?`${units.length} distinct overlap ${units.length===1?'unit':'units'} · ${r.relationship}`:'No coded overlap in this dimension';
    detailBadge.textContent=`${intensityLabel(r.intensity)} directness`;
    detailBadge.style.opacity=r.intensity?1:.5;
    renderUnitCards(units);
    currentPrograms=getProgramsForUnits(units);
    renderPrograms(); renderBenchmarks(units); updateSelectedBubble();
    history.replaceState(null,'',`#details/${r.diseaseSlug}/${r.factorSlug}`);
    if(scroll) document.getElementById('details').scrollIntoView({behavior:'smooth',block:'start'});
    window.setTimeout(hideTooltip, scroll ? 650 : 0);
  }
  function initialCell(){
    const parts=location.hash.split('/');
    if(parts[0]==='#details'&&parts.length>=3){
      const id=`${parts[1]}--${parts[2]}`;
      if(D.plot.some(p=>p.cellId===id&&p.overlapCount>0))return id;
    }
    return D.plot.find(p=>p.diseaseArea==='Solid-tumor oncology'&&p.factor==='Shared indications').cellId;
  }
  function renderSources(){
    const list=document.getElementById('source-list');
    list.innerHTML=D.sources.map(s=>`<li><div><a class="source-name" href="${esc(s.url)}" target="_blank" rel="noopener">${esc(s.name)} ↗</a></div><div>${esc(s.date)}</div><div class="source-use">${esc(s.use)}</div></li>`).join('');
  }
  searchInput.addEventListener('input',renderPrograms);
  renderChart(); renderSources(); selectCell(initialCell(),false);
})();
'''
(OUT / 'assets' / 'app.js').write_text(app_js, encoding='utf-8')

# -----------------------------
# HTML
# -----------------------------
index_html = r'''<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <meta name="description" content="Interactive AstraZeneca and Bristol Myers Squibb late-stage portfolio overlap analysis with indication-level evidence and trial details.">
  <meta property="og:title" content="AstraZeneca + BMS: Portfolio Fit or Collision?">
  <meta property="og:description" content="Click each overlap bubble to inspect the exact indications, modalities, mechanisms and programs behind the count.">
  <meta property="og:type" content="website">
  <meta property="og:image" content="assets/figure-fallback.png">
  <title>AstraZeneca + BMS: Portfolio Fit or Collision?</title>
  <link rel="icon" href="assets/favicon.svg" type="image/svg+xml">
  <link rel="stylesheet" href="assets/styles.css">
</head>
<body>
<a class="skip-link" href="#main">Skip to content</a>
<header class="site-header">
  <div class="nav-wrap">
    <div class="brand"><span class="brand-mark" aria-hidden="true"></span><span>Portfolio collision explorer</span></div>
    <nav class="nav-links" aria-label="Primary navigation">
      <a href="#chart-section">Interactive chart</a><a href="#details">Evidence</a><a href="#methodology">Methodology</a><a href="#downloads">Downloads</a>
    </nav>
  </div>
</header>
<main id="main">
  <section class="hero">
    <div class="container">
      <div class="eyebrow">Interactive portfolio scenario · matched February 2026 snapshots</div>
      <h1>AstraZeneca + BMS:<br>where does the portfolio <span class="highlight">really collide?</span></h1>
      <p class="hero-copy">The headline overlap is oncology, but not every shared area is equally direct. This explorer moves from portfolio counts to the exact indication, modality, target and stage evidence behind each bubble.</p>
      <div class="hero-actions"><a class="button primary" href="#chart-section">Explore the chart</a><a class="button secondary" href="downloads/AZ_BMS_Portfolio_Interactive_Website_Backup.xlsx">Download workbook</a></div>
      <div class="status-note"><span class="status-dot" aria-hidden="true"></span><span>This is a portfolio thought experiment, not a claim that a transaction is proceeding. Clinical and commercial data are sourced; launch windows are clearly labeled analyst scenarios.</span></div>
    </div>
  </section>

  <section class="section" id="chart-section">
    <div class="container">
      <div class="section-head"><div><div class="kicker">Main figure</div><h2>Click a bubble to see what is underneath it</h2><p>Top bars retain the late-stage program footprint. Bubble size and label show the number of distinct overlap units; red tone shows how directly the portfolios collide.</p></div></div>
      <div class="card chart-card">
        <div class="chart-toolbar"><div><div class="chart-title">Late-stage footprint + overlap matrix</div><div class="chart-subtitle">Disease areas are columns; portfolio dimensions are rows</div></div><div class="chart-hint">Click or press Enter on a red bubble</div></div>
        <div class="chart-scroll"><div class="chart-wrap" id="chart"></div></div>
        <noscript><img src="assets/figure-fallback.png" alt="Static portfolio collision chart" style="max-width:100%;height:auto"></noscript>
        <div class="chart-caption"><div class="legend-line"><span class="legend-company"><span class="legend-swatch az"></span>AstraZeneca</span><span class="legend-company"><span class="legend-swatch bms"></span>BMS</span></div><div>Counts are distinct shared territories or classes, not every possible pair of assets.</div></div>
      </div>
      <div class="insight-strip"><div class="insight-label">What the chart says</div><div><strong>Solid tumors and hematology contain the deepest collision, but the overlap ranges from exact indication or target competition to broad platform adjacency.</strong><p>The evidence section separates those cases and shows the programs, clinical setting, representative trial timing, illustrative launch window and commercial context.</p></div></div>
    </div>
  </section>

  <section class="section" id="details">
    <div class="container">
      <div class="section-head"><div><div class="kicker">Bubble evidence</div><h2>From count to exact overlap</h2><p>Select an overlap unit card to narrow the table to one indication, modality, mechanism or stage band.</p></div></div>
      <div class="card detail-card">
        <div class="selection-head"><div><div class="kicker">Selected chart cell</div><h3 id="detail-title">Loading…</h3><div class="selection-meta" id="detail-meta"></div></div><div class="selection-badge" id="detail-badge"></div></div>
        <div class="unit-grid" id="unit-grid"></div>
        <div class="detail-controls"><h4>Relevant program detail</h4><input class="search" id="program-search" type="search" placeholder="Search programs, stages or settings" aria-label="Search program details"></div>
        <div class="table-wrap">
          <table class="data-table">
            <thead><tr><th>Company</th><th>Program / modality</th><th>Stage</th><th>Representative trial</th><th>Potential launch indication</th><th>Clinical setting / line</th><th>Primary completion</th><th>Illustrative launch window</th><th>Historical sales</th><th>Source / note</th></tr></thead>
            <tbody id="program-table-body"></tbody>
          </table>
        </div>
        <div class="commercial-section" id="commercial-section"><h4>Historical / commercial scale context</h4><div class="benchmark-grid" id="benchmark-grid"></div></div>
      </div>
    </div>
  </section>

  <section class="section" id="methodology">
    <div class="container">
      <div class="section-head"><div><div class="kicker">Methodology</div><h2>How to read the analysis</h2></div></div>
      <div class="method-grid">
        <div class="card method-card"><div class="number">1</div><h3>Late-stage program</h3><p>A unique lead investigational asset in Phase II or Phase III/pivotal/registration development. Approved-indication expansions and major lifecycle-management programs are excluded from the core count.</p></div>
        <div class="card method-card"><div class="number">2</div><h3>Overlap count</h3><p>One count equals one distinct shared indication territory, modality class, target/mechanism theme or stage band. It is not the number of all possible AZ-BMS asset pairs.</p></div>
        <div class="card method-card"><div class="number">3</div><h3>Launch and sales fields</h3><p>Launch windows are illustrative analyst scenarios based on posted trial timing and stage-specific lags. Investigational assets have no historical sales; company-reported brand revenue is shown only for scale context.</p></div>
      </div>
    </div>
  </section>

  <section class="section" id="downloads">
    <div class="container">
      <div class="section-head"><div><div class="kicker">Reproducibility package</div><h2>Data, workbook and static exports</h2><p>The site has no build dependency. All chart and detail data are also available as CSV and JSON.</p></div></div>
      <div class="download-grid">
        <a class="card download-card" href="downloads/AZ_BMS_Portfolio_Interactive_Website_Backup.xlsx"><div class="download-type">XLSX</div><h3>Backup workbook</h3><p>Asset data, overlap units, formulas, trial details and commercial benchmarks.</p></a>
        <a class="card download-card" href="data/portfolio_data.json"><div class="download-type">JSON</div><h3>Website data</h3><p>Single structured export used by the interactive application.</p></a>
        <a class="card download-card" href="data/program_details.csv"><div class="download-type">CSV</div><h3>Program details</h3><p>Trial stage, setting, timing, launch scenario and source URLs.</p></a>
        <a class="card download-card" href="data/overlap_units.csv"><div class="download-type">CSV</div><h3>Overlap evidence</h3><p>Exact indications, modalities, mechanisms and stage bands behind the bubbles.</p></a>
        <a class="card download-card" href="downloads/AZ_BMS_Portfolio_Collision_Static.pdf"><div class="download-type">PDF</div><h3>Static figure</h3><p>Print-ready version of the original chart.</p></a>
        <a class="card download-card" href="downloads/build_az_bms_lollipop.R"><div class="download-type">R</div><h3>Chart script</h3><p>Reproducible ggplot2 implementation for the static figure.</p></a>
      </div>
    </div>
  </section>

  <section class="section" id="sources">
    <div class="container">
      <div class="section-head"><div><div class="kicker">Primary sources</div><h2>Source trail</h2><p>Individual ClinicalTrials.gov links appear in the detailed program table.</p></div></div>
      <div class="card sources"><ul class="source-list" id="source-list"></ul></div>
    </div>
  </section>
</main>
<footer><div class="container"><div class="footer-grid"><div><strong>Portfolio collision explorer</strong><br>Data as of 5 August 2026; matched pipeline counts use February 2026 company snapshots.</div><div>For strategic discussion only. Not medical, legal or investment advice.</div></div></div></footer>
<div class="tooltip" id="tooltip" role="tooltip"></div>
<script src="assets/data.js"></script><script src="assets/app.js"></script>
</body>
</html>
'''
(OUT / 'index.html').write_text(index_html, encoding='utf-8')
(OUT / '.nojekyll').write_text('', encoding='utf-8')

# README for GitHub hosting
readme = '''# AstraZeneca + BMS portfolio collision explorer

A static, dependency-free website package for GitHub Pages. The main chart reproduces the stacked late-stage portfolio bars and interactive overlap bubbles. Clicking a bubble scrolls to the evidence table and exposes the exact indication, modality, target/mechanism or stage units behind the count.

## Publish on GitHub Pages

1. Create a new GitHub repository.
2. Upload **all files and folders from this package root** to the repository root.
3. In GitHub, open **Settings → Pages**.
4. Under **Build and deployment**, choose **Deploy from a branch**.
5. Select `main` and `/ (root)`, then save.

The site uses plain HTML, CSS, JavaScript and local data files. No npm installation or build step is required.

## Local preview

From the package folder:

```bash
python -m http.server 8000
```

Then open `http://localhost:8000`.

## Update the data

- `assets/data.js` is the browser-ready data file.
- `data/portfolio_data.json` is the same dataset as JSON.
- CSV exports are in `data/`.
- The source workbook is in `downloads/`.

After changing the data model, keep `assets/data.js` and `data/portfolio_data.json` synchronized. The included `build_az_bms_website.py` generator in the parent deliverable can be used as the authoritative build path.

## Analytical definitions

- **Late-stage program:** unique lead investigational asset in Phase II or Phase III/pivotal/registration development; approved-indication expansions and major lifecycle-management programs excluded from the core count.
- **Overlap count:** distinct shared territory/class, not all possible asset-pair permutations.
- **Launch window:** illustrative analyst scenario from trial primary-completion timing plus a stage-specific lag; not company guidance.
- **Historical sales:** investigational assets have no historical sales. Company-reported brand revenue is used only as contextual scale and is generally not indication-specific.

## License and trademarks

The website code is provided under the MIT License. AstraZeneca and Bristol Myers Squibb names are used descriptively; all trademarks remain the property of their respective owners. The package does not include company logo artwork.
'''
(OUT / 'README.md').write_text(readme, encoding='utf-8')

license_text = '''MIT License

Copyright (c) 2026

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
'''
(OUT / 'LICENSE').write_text(license_text, encoding='utf-8')

# Data dictionary
(OUT / 'DATA_DICTIONARY.md').write_text('''# Data dictionary

## `chart_data.csv`
One row per disease-area × overlap-factor chart cell.

- `overlapCount`: distinct overlap units represented by bubble size/label.
- `intensity`: analyst-coded directness from 0 to 3.
- `azPrograms`, `bmsPrograms`: late-stage program counts in the stacked bar.

## `overlap_units.csv`
One row per exact shared territory/class. This is the evidence behind each bubble count.

## `program_details.csv`
One row per program in the curated trial-level detail set, plus pipeline-snapshot fields for the broader mapped portfolio.

- `potentialLaunchIndication`: potential first indication inferred from the representative study, not company guidance.
- `illustrativeLaunchWindow`: trial primary-completion year plus a stage-specific lag.
- `historicalSales`: investigational assets are marked as not launched.

## `commercial_benchmarks.csv`
Company-reported brand revenue used only to illustrate franchise scale. Figures are generally worldwide and not indication-specific.
''', encoding='utf-8')

# Add the generator itself to package for transparent reproducibility.
shutil.copy2(Path(__file__), OUT / 'build_az_bms_website.py')

# Zip the entire package.
if ZIP_PATH.exists(): ZIP_PATH.unlink()
with zipfile.ZipFile(ZIP_PATH, 'w', compression=zipfile.ZIP_DEFLATED, compresslevel=9) as z:
    for path in sorted(OUT.rglob('*')):
        if path.is_file():
            z.write(path, arcname=path.relative_to(OUT))

print(json.dumps({
    'site_dir': str(OUT),
    'zip': str(ZIP_PATH),
    'files': len([p for p in OUT.rglob('*') if p.is_file()]),
    'programs': len(enriched_programs),
    'curated_programs': len(curated_programs),
    'overlap_units': len(overlap_units),
    'plot_cells': len(plot_rows),
}, indent=2))
